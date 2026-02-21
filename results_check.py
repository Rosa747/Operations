import pandas as pd
from openpyxl import load_workbook


aircrafts_arr = pd.read_excel('flight_schedule_test.xlsx', sheet_name='A', header = 0)
aircrafts_dep = pd.read_excel('flight_schedule_test.xlsx', sheet_name='D', header = 0)
aircrafts = pd.concat([aircrafts_arr, aircrafts_dep], ignore_index=True)["id"]

route_times_df = pd.DataFrame()



def read_solution_sheet(path, sheet_name):
    # 1) Read A1 ("Chosen route: ...")
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    a1 = ws["A1"].value or ""

    # Extract route name after "Chosen route:"
    chosen_route = a1.split(":", 1)[1].strip() if ":" in a1 else None

    # 2) Read Node/Time table (header on row 2)
    df = pd.read_excel(path, sheet_name=sheet_name, header=1)  # row 2 is header
    df = df[["Node", "Time"]].dropna(subset=["Node"])          # keep valid rows

    # Convert to dict: node -> time
    node_times = dict(zip(df["Node"].astype(str), df["Time"].astype(float)))

    return {"chosen_route": chosen_route, "times": node_times}

nodes_times_df = pd.DataFrame()

for aircraft_id in aircrafts:
    data = read_solution_sheet("solution_output.xlsx", sheet_name=aircraft_id)

    route_dict = data["times"]
    nodes_times_df = pd.concat(
        [nodes_times_df, pd.DataFrame([route_dict], index=[aircraft_id])],
        sort=False,
    )
    nodes_times_df.loc[aircraft_id, "chosen_route"] = data["chosen_route"]

# make chosen_route the first column
cols = ["chosen_route"] + [col for col in nodes_times_df.columns if col != "chosen_route"]
nodes_times_df = nodes_times_df[cols]
print(nodes_times_df)

# Nieuwe route met gate onderaan voor om taxien in plaats van langzamer
